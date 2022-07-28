def main():
    # Учёт техники в компании, ключ это предмет в компании, значение это список["адрес","человек","id","цена"]
    # вывод информации, перемещение в другое место, перепривяку к другому человеку, удаление, добавление, редактирование
    # сделать словарь для учёта офисов
    import openpyxl as xl
    from openpyxl.styles import Border,Side
    # Отрытие обрабатываемой страницы .xlsx файла
    techs=xl.open('D:\Python\Accounting.xlsx',data_only=True)
    sheet = techs['техника']
    offices={"Офис_1":"ул. Одинцова 8",'Офис_2':'ул. Кульман 9','Офис_3':'ул. Мележа, 1'}
    exit = True
    while exit:
        choose=int(input("1.Заниматься предметами\n2.Заниматься офисами\n3.ВЫХОД\n-> "))
        if choose ==1:
            while exit:
                choose = int(input("1.ДОБАВИТЬ\n2.РЕДАКТИРОВАТЬ\n3.УДАЛИТЬ\n4.ПЕРЕМЕЩЕНИЕ\n5.ПЕРЕПРИВЯЗКА\n"
                                   "6.ВЫВОД ВСЕХ\n7.ВЫВОД ТЕХНИКИ В ОФИСЕ\n8.ВЫХОД\n-> "))
                if choose == 1:         # ДОБАВЛЕНИЕ НОВОЙ ЗАПИСИ
                    name = input("Введите название новой техники: ")
                    adres = input("Введите офис новой техники: ")
                    owner = input("Введите ответственного владельца: ")
                    coast= int(input("Введите цену новой техники: "))
                    # Поиск первой свободной строки от записей и определение id новой техники
                    for r in range(3,sheet.max_row):
                        if sheet['A'][r].value is None:
                            text_row=int(sheet['A'][r-1].value)+1
                            new_id= r+1
                            break
                    new_data=(str(text_row),name,offices[adres],adres,owner,str(coast))
                    # запись в .xlsx файл новой строки и обрисовка границ ячейки под стиль файла
                    for row in range(0,len(new_data)):
                        sheet[new_id][row].value=new_data[row]
                        bord=Border(top=Side(border_style='thin'),
                                    bottom=Side(border_style='thin'),
                                    left=Side(border_style='thin'),
                                    right=Side(border_style='thin'))
                        sheet[new_id][row].border=bord
                    techs.save('D:\Python\Accounting.xlsx')         # сохранение изменений в файле
                elif choose == 2:           #  РЕДАКТИРОВАНИЕ ЗАПИСИ
                    name = input("Введите название техники: ")
                    of=input("Введите офис: ")
                    # Поиск строки по технике и офису в заполненных строках и запись новых данных
                    for r in range(4,sheet.max_row):
                        if sheet['A'][r].value!=None:
                            if sheet['B'][r].value==name and sheet['D'][r].value==of:
                                new_coast=int(input("Введите новую цену: "))
                                sheet['F'][r].value=new_coast
                                techs.save('D:\Python\Accounting.xlsx')
                                break
                    else:
                        print("Нет такой техники!")
                elif choose == 3:       #  УДАЛЕНИЕ СТРОКИ
                    name =input("Введите название техники: ")
                    of = input("Введите офис: ")
                    # Поиск строки по технике и офису в заполненных строках
                    for row in range(3, sheet.max_row + 1):
                        if sheet[row][0].value != None:
                            if str(sheet['B'][row].value)==name and sheet['D'][row].value==of:
                    # Удаление найденной строки
                                sheet.delete_rows(idx=row, amount=1)
                                techs.save('D:\Python\Accounting.xlsx')
                                break
                    else:
                        print("Нет такой техники")
                elif choose == 4:       # ПЕРЕМЕЩЕНИЕ ТЕХНИКИ В ДРУГОЙ ОФИС
                    name = input("Введите технику: ")
                    of = input("Введите офис: ")
                    for row in range(3, sheet.max_row + 1):
                        if sheet[row][0].value != None:
                            if sheet['B'][row].value==name and sheet['D'][row].value==of:
                                new_adres = input("Введите новый офис: ")
                                new_person = input("Введите нового ответственного: ")
                                sheet['D'][row].value = new_adres
                                sheet['C'][row].value=offices[new_adres]       # запись нового адреса по офису
                                sheet['E'][row].value = new_person
                                techs.save('D:\Python\Accounting.xlsx')
                                break
                    else:
                        print("Нет такой техники")
                elif choose == 5:       # ПЕРЕПРИВЯЗКА ТЕХНИКИ К НОВОМУ ЧЕЛОВЕКУ
                    name = input("Введите технику: ")
                    pers = input("Введите ответственного: ")
                    for row in range(3, sheet.max_row + 1):
                        if sheet[row][0].value != None:
                            if sheet['B'][row].value == name and sheet['E'][row].value == pers:
                                new_person = input("Введите нового ответственного: ")
                                sheet['E'][row].value = new_person
                                techs.save('D:\Python\Accounting.xlsx')
                                break
                    else:
                        print("Нет такого товара")
                elif choose == 6:       #  ВЫВОД НА ЭКРАН ВСЕХ ЗАПОЛНЕННЫХ СТРОК ФАЙЛА
                    for row in range(3, sheet.max_row + 1):
                        if sheet[row][0].value != None:
                            print(sheet[row][0].value,
                              (sheet[row][1].value).center(18),     # центровка данных при выводе на экран
                              (sheet[row][2].value).center(18),
                              (sheet[row][3].value).center(18),
                              (sheet[row][4].value).center(18),
                              str(sheet[row][5].value).center(18))
                elif choose==7:     #   ВЫВОД НА ЭКРАН ТЕХНИКИ В КОНКРЕТНОМ ОФИСЕ
                    name = input("Введите офис: ")
                    if name in offices.keys():
                        # вывод шапки файла на экран
                        print(sheet['A3'].value,
                              (sheet['B3'].value).center(18),
                              (sheet['C3'].value).center(18),
                              (sheet['D3'].value).center(18),
                              (sheet['E3'].value).center(18),
                              str(sheet['F3'].value).center(18))
                        # вывод на экран искомой информации
                        for element in range(3,sheet.max_row+1):
                            if sheet[element][0].value != None:
                                if sheet[element][3].value==name:
                                    print(sheet[element][0].value,
                                         (sheet[element][1].value).center(18),
                                         (sheet[element][2].value).center(18),
                                         (sheet[element][3].value).center(18),
                                         (sheet[element][4].value).center(18),
                                         str(sheet[element][5].value).center(18))
                    else:
                        print("Нет такого офиса")
                elif choose == 8:
                    exit = False
                else:
                    print("Некорректный ввод, повторите попытку!")
            exit = True
        elif choose ==2:
            while exit:
                choose = int(input("1.ДОБАВИТЬ\n2.РЕДАКТИРОВАТЬ\n3.УДАЛИТЬ\n4.ВЫВОД ВСЕХ\n5.ВЫХОД\n-> "))
                if choose == 1:
                    name =input("Ввести название офиса: ")
                    adres = input("Ввести адрес: ")
                    # Поиск первой свободной строки от записей и определение id новой техники
                    for r in range(3, sheet.max_row):
                        if sheet['D'][r].value == None:
                            sheet['D'][r].value == name
                            sheet['C'][r].value = adres
                            techs.save('D:\Python\Accounting.xlsx')  # сохранение изменений в файле
                            break
                elif choose == 2:    #  РЕДАКТИРОВАНИЕ ЗАПИСИ
                    name = input("Введите офис: ")
                    new_adres = input("Ввести адрес: ")
                    # Поиск строки по офису в заполненных строках и запись новых данных
                    for r in range(4, sheet.max_row):
                        if sheet['D'][r].value != None:
                            if sheet['D'][r].value == name:
                                sheet['C'][r].value = new_adres
                                techs.save('D:\Python\Accounting.xlsx')
                                break
                    else:
                        print("Нет такого офиса")
                elif choose == 3:   #  УДАЛЕНИЕ СТРОКИ
                    try:
                        offis = input("Введите офис: ")
                        # Поиск строки по офису в заполненных строках
                        for row in range(3, sheet.max_row + 1):
                            if sheet[row][2].value != None:
                                if str(sheet['D'][row].value) == offis:
                        # Удаление найденной строки
                                    sheet['D'][row].value = "None"
                                    techs.save('D:\Python\Accounting.xlsx')
                                    break
                    except AttributeError:
                        print("Ничего страшного")
                    else:
                        print("Нет такого офиса")
                elif choose == 4:
                    for row in range(3, sheet.max_row + 1):
                        if sheet[row][3].value != None:
                                print ((sheet[row][3].value).center(18),
                                       (sheet[row][2].value).center(18))
                elif choose == 5:
                    exit = False
                else:
                    print("Некорректный ввод, повторите попытку")
            exit = True
        elif choose == 3:
            techs.close()       # закрытие файла
            exit = False
        else:
            print("Некорректный ввод, повторите попытку")

main()